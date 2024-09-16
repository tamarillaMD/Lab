import os
import pandas as pd
from openpyxl import load_workbook

# Define the directory containing the Excel files
folder_path = r"C:\Users\nshcher\Desktop\Pivot_test\test_R_for_DB"

# Define the universal column renaming using partial matching
universal_rename = {
    'Genotype': 'genotype', 
    'Quality': 'quality',
    'Alternative depth': 'genotype', 
    'Total depth': 'quality',
    'Alternate allele fraction': 'Alternate allele fraction'
}

# Excel formula to be added
excel_formula = (
    '=IF(XLOOKUP("*"&[@[Genome-ID]]&"*",All_Patients[ID],All_Patients[KL Lab ID],"Not found",2)="Not found",XLOOKUP([@[Genome-ID]],All_Patients[ID],All_Patients[KL Lab ID],"Not found",2),XLOOKUP("*"&[@[Genome-ID]]&"*",All_Patients[ID],All_Patients[KL Lab ID],"Not found",2))'
    )

# Iterate through all Excel files in the folder
for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx"):  # Ensure we're working with Excel files
        file_path = os.path.join(folder_path, file_name)
        
        # Load the Excel file
        df = pd.read_excel(file_path)
        
        # Retain first 65 columns (up to 'BM')
        columns_to_keep = df.columns[:65].tolist()

        # Identify the columns related to the file name
        file_base_name = os.path.splitext(file_name)[0]  # Extract the file name without .xlsx
        file_related_columns = [col for col in df.columns if file_base_name in col]
        
        # Retain only necessary columns and add them to the list
        columns_to_keep.extend(file_related_columns)

        # Subset the dataframe
        df = df[columns_to_keep]
        
        # Rename columns using partial string matching
        for old_col in df.columns:
            for key, new_col in universal_rename.items():
                if key in old_col:  # Check if the keyword is in the column name
                    df = df.rename(columns={old_col: new_col})
        
        # Save the dataframe to a new Excel file with openpyxl
        temp_file_path = os.path.join(folder_path, "temp_" + file_name)
        df.to_excel(temp_file_path, index=False)

        # Load the workbook and select the active worksheet
        wb = load_workbook(temp_file_path)
        ws = wb.active

        # Insert the file name as the first column
        ws.insert_cols(1)
        ws['A1'] = 'Genome-ID'
        for row in range(2, ws.max_row + 1):
            ws[f'A{row}'] = file_base_name
        
        # Insert the formula as the second column
        ws.insert_cols(2)
        ws['B1'] = 'formula_column'
        for row in range(2, ws.max_row + 1):
            ws[f'B{row}'] = excel_formula
        
        # Save the modified file (overwrite the original or save to a new file)
        wb.save(file_path)
        
        # Optionally, remove the temporary file
        os.remove(temp_file_path)

print("Processing complete.")