import os
import pandas as pd
from openpyxl import load_workbook

# Define the path for the folder with the Excel files and the Excel file that contains the renaming table
folder_path = r"C:\Users\nshcher\Desktop\241009_allcopied_allchanged_testing_with duplicates_overwrite"
renaming_table_path = r"C:\Users\nshcher\Desktop\renaming_table.xlsx"

# Load the renaming table
renaming_df = pd.read_excel(renaming_table_path)

# Iterate over the files in the folder and rename according to the table
for _, row in renaming_df.iterrows():
    old_name = row['VarFish ID']
    new_name = row['real Org ID']
    old_file_path = os.path.join(folder_path, old_name + ".xlsx")
    new_file_path = os.path.join(folder_path, new_name + ".xlsx")
    
    if os.path.exists(old_file_path):
        # Check if the file with the new name already exists, if so, skip renaming
        if not os.path.exists(new_file_path):
            # Rename the file
            os.rename(old_file_path, new_file_path)
            print(f"Renamed {old_name} to {new_name}")
            
            # Load the renamed Excel file
            wb = load_workbook(new_file_path)
            
            # Iterate through each sheet and update the first column
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.cell(row=1, column=1, value=new_name)  # Update the first column, first row with new name
                
            # Save the workbook after renaming inside the sheets
            wb.save(new_file_path)
            print(f"Updated first column with {new_name} in all sheets of {new_file_path}")
        else:
            print(f"File {new_name}.xlsx already exists. Skipping renaming.")
    else:
        print(f"File {old_name}.xlsx not found in the folder.")
        
print("Renaming process completed.")